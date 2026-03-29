import os

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_worksheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)


def add_price_chart(ws_chart, ws_data, num_rows: int, company_name: str):
    chart = LineChart()
    chart.title = f"{company_name} Price Trend"
    chart.y_axis.title = "Close Price"
    chart.x_axis.title = "Date"

    data = Reference(ws_data, min_col=5, min_row=1, max_row=num_rows)
    cats = Reference(ws_data, min_col=1, min_row=2, max_row=num_rows)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18

    ws_chart.add_chart(chart, "B2")


def add_financial_chart(ws_chart, ws_data, num_rows: int, company_name: str):
    chart = BarChart()
    chart.title = f"{company_name} Revenue vs PAT"
    chart.y_axis.title = "INR Crore"
    chart.x_axis.title = "Year"

    data = Reference(ws_data, min_col=5, max_col=7, min_row=1, max_row=num_rows)
    cats = Reference(ws_data, min_col=4, min_row=2, max_row=num_rows)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 18

    ws_chart.add_chart(chart, "B18")


def export_company_report(
    company_name: str,
    price_df: pd.DataFrame,
    metrics: dict,
    narrative_lines: list[str],
    company_financials_df: pd.DataFrame | None = None,
    output_folder: str = "reports",
) -> str:
    os.makedirs(output_folder, exist_ok=True)

    safe_name = company_name.lower().replace(" ", "_")
    output_path = os.path.join(output_folder, f"{safe_name}_report.xlsx")

    summary_df = pd.DataFrame([metrics])
    narrative_df = pd.DataFrame({"Narrative": narrative_lines})

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        price_df.to_excel(writer, sheet_name="Price Data", index=False)
        narrative_df.to_excel(writer, sheet_name="Narrative", index=False)

        if company_financials_df is not None and not company_financials_df.empty:
            company_financials_df.to_excel(writer, sheet_name="Financials", index=False)

        workbook = writer.book
        workbook.create_sheet("Charts")

        ws_summary = workbook["Summary"]
        ws_price = workbook["Price Data"]
        ws_narrative = workbook["Narrative"]
        ws_chart = workbook["Charts"]

        style_worksheet(ws_summary)
        style_worksheet(ws_price)
        style_worksheet(ws_narrative)

        add_price_chart(ws_chart, ws_price, len(price_df) + 1, company_name)

        if company_financials_df is not None and not company_financials_df.empty:
            ws_financials = workbook["Financials"]
            style_worksheet(ws_financials)
            add_financial_chart(ws_chart, ws_financials, len(company_financials_df) + 1, company_name)

    return output_path


def export_combined_summary(
    summary_df: pd.DataFrame,
    narrative_lines: list[str],
    latest_financial_snapshot_df: pd.DataFrame | None = None,
    financial_narrative_df: pd.DataFrame | None = None,
    output_folder: str = "reports",
) -> str:
    os.makedirs(output_folder, exist_ok=True)

    output_path = os.path.join(output_folder, "combined_company_summary.xlsx")

    rankings_df = summary_df.sort_values("5Y Return (%)", ascending=False).reset_index(drop=True)
    rankings_df.insert(0, "Rank", range(1, len(rankings_df) + 1))

    narrative_df = pd.DataFrame({"Narrative": narrative_lines})

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Company Summary", index=False)
        rankings_df.to_excel(writer, sheet_name="Rankings", index=False)
        narrative_df.to_excel(writer, sheet_name="Market Narrative", index=False)

        if latest_financial_snapshot_df is not None and not latest_financial_snapshot_df.empty:
            latest_financial_snapshot_df.to_excel(writer, sheet_name="Financial Snapshot", index=False)

        if financial_narrative_df is not None and not financial_narrative_df.empty:
            financial_narrative_df.to_excel(writer, sheet_name="Financial Narrative", index=False)

        workbook = writer.book

        for sheet_name in workbook.sheetnames:
            style_worksheet(workbook[sheet_name])

        ws_rankings = workbook["Rankings"]

        return_col_idx = None
        for idx, cell in enumerate(ws_rankings[1], start=1):
            if cell.value == "5Y Return (%)":
                return_col_idx = idx
                break

        if return_col_idx:
            col_letter = get_column_letter(return_col_idx)
            ws_rankings.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{ws_rankings.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="63BE7B",
                ),
            )

    return output_path
